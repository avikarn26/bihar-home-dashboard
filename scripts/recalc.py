#!/usr/bin/env python3
import sys
import json
import subprocess
import tempfile
import os
from pathlib import Path

def recalc_excel(file_path, timeout=30):
    try:
        file_path = str(Path(file_path).resolve())
        cmd = [
            'soffice',
            '--headless',
            '--convert-to', 'xlsx',
            '--outdir', tempfile.gettempdir(),
            file_path
        ]
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)

        if result.returncode != 0:
            return {"status": "error", "message": result.stderr}

        from openpyxl import load_workbook
        wb = load_workbook(file_path, data_only=False)
        error_summary = {}
        total_errors = 0
        total_formulas = 0

        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.data_type == 'f':
                        total_formulas += 1
                        if isinstance(cell.value, str) and cell.value.startswith('='):
                            try:
                                val = cell.value
                                if '#' in str(cell.value):
                                    error_type = 'unknown'
                                    if '#REF!' in val: error_type = '#REF!'
                                    elif '#DIV/0!' in val: error_type = '#DIV/0!'
                                    elif '#VALUE!' in val: error_type = '#VALUE!'
                                    elif '#N/A' in val: error_type = '#N/A'

                                    if error_type not in error_summary:
                                        error_summary[error_type] = {'count': 0, 'locations': []}
                                    error_summary[error_type]['count'] += 1
                                    error_summary[error_type]['locations'].append(f"{sheet}!{cell.coordinate}")
                                    total_errors += 1
                            except:
                                pass

        return {
            "status": "success" if total_errors == 0 else "errors_found",
            "total_errors": total_errors,
            "total_formulas": total_formulas,
            "error_summary": error_summary if error_summary else {}
        }
    except Exception as e:
        return {"status": "error", "message": str(e)}

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(json.dumps({"status": "error", "message": "Usage: recalc.py <file.xlsx> [timeout]"}))
        sys.exit(1)

    file_path = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 30
    result = recalc_excel(file_path, timeout)
    print(json.dumps(result))
