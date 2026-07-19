#!/usr/bin/env python3
"""Build dashboard_data.json + index.html embedded JSON from Excel (single source of truth).

Usage:  python3 scripts/build_dashboard.py [--check]
  --check : compute and reconcile only, don't write files.

Everything numeric is COMPUTED from data/Project_Home_Tracker.xlsx:
  - All Transactions  -> billed totals, vendors, categories, monthly, vendor tx lists
  - Payment Tracker   -> paid totals, payers, payments list, vendor paid, unpaid dues
Non-derivable display data (meta/phase/notes, brickAdvance, materials, vendor
display types, colors) is carried over from the existing dashboard_data.json.
"""
import json
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "data" / "Project_Home_Tracker.xlsx"
JSON_PATH = ROOT / "dashboard_data.json"
HTML_PATH = ROOT / "index.html"

# Excel category -> dashboard category
CATEGORY_MAP = {
    "Building Materials": "Building Materials",
    "Steel/Iron": "Steel / Iron",
    "Steel / Iron": "Steel / Iron",
    "Professional Fees": "Professional Fees",
    "Land Development": "Land Development",
    "Labour": "Labour",
    "Transport": "Transport · Food · Misc",
    "Food/Provisions": "Transport · Food · Misc",
    "Miscellaneous": "Transport · Food · Misc",
    "Transport · Food · Misc": "Transport · Food · Misc",
}

# Payment Tracker "Paid By" -> dashboard payer
PAYER_MAP = {
    "Self": "Avinash",
    "Self-GPay": "Avinash",
    "Self-Bank Transfer": "Avinash",
    "Papa": "Papa",
    "Chota Bhai": "Chota Bhai",
}

BRICK_VENDOR = "Brick Company"  # advance paid > billed by design; never "due"


def cell_amount(ws, r):
    """Net amount for a tx row: gross (cached L or qty*rate) minus discount % (col M)."""
    gross = ws.cell(r, 12).value
    if not isinstance(gross, (int, float)):
        qty, rate = ws.cell(r, 9).value, ws.cell(r, 11).value
        if isinstance(qty, (int, float)) and isinstance(rate, (int, float)):
            gross = float(qty) * float(rate)
        else:
            return None
    disc = ws.cell(r, 13).value
    if isinstance(disc, (int, float)) and 0 < disc < 1:
        gross = gross * (1 - disc)
    return float(gross)


def read_transactions(wb):
    ws = wb["All Transactions"]
    rows = []
    for r in range(5, ws.max_row + 1):
        if ws.cell(r, 1).value == "TOTALS":
            break
        d = ws.cell(r, 2).value
        vendor = ws.cell(r, 8).value
        if not isinstance(d, datetime) or not vendor:
            continue
        amt = cell_amount(ws, r)
        if amt is None:
            continue
        rows.append({
            "date": d,
            "category": str(ws.cell(r, 5).value or "").strip(),
            "sub": str(ws.cell(r, 6).value or "").strip(),
            "desc": str(ws.cell(r, 7).value or "").strip(),
            "vendor": str(vendor).strip(),
            "amount": round(amt),
        })
    return rows


def read_payments(wb):
    ws = wb["Payment Tracker"]
    rows = []
    for r in range(5, ws.max_row + 1):
        d = ws.cell(r, 2).value
        amt = ws.cell(r, 7).value
        if not isinstance(d, datetime) or not isinstance(amt, (int, float)):
            continue
        rows.append({
            "date": d,
            "by": str(ws.cell(r, 3).value or "").strip(),
            "to": str(ws.cell(r, 4).value or "").strip(),
            "desc": str(ws.cell(r, 6).value or "").strip(),
            "amount": round(float(amt)),
            "mode": str(ws.cell(r, 8).value or "").strip(),
        })
    return rows


def build(d_old, txs, pays):
    d = json.loads(json.dumps(d_old))  # deep copy; carry over display-only data

    total_billed = sum(t["amount"] for t in txs)
    total_paid = sum(p["amount"] for p in pays)

    # --- vendors ---
    billed_by_vendor = defaultdict(int)
    tx_by_vendor = defaultdict(list)
    for t in txs:
        billed_by_vendor[t["vendor"]] += t["amount"]
        tx_by_vendor[t["vendor"]].append([t["date"].strftime("%d-%b"), t["desc"], t["amount"]])
    paid_by_vendor = defaultdict(int)
    for p in pays:
        paid_by_vendor[p["to"]] += p["amount"]

    old_vendors = {v["name"]: v for v in d_old["vendors"]}
    vendors = []
    all_names = list(dict.fromkeys(list(billed_by_vendor) + list(paid_by_vendor)))
    # keep existing dashboard ordering first
    ordered = [n for n in old_vendors if n in all_names] + [n for n in all_names if n not in old_vendors]
    for name in ordered:
        old = old_vendors.get(name, {})
        billed = billed_by_vendor.get(name, 0)
        paid = paid_by_vendor.get(name, 0)
        if name == BRICK_VENDOR:
            status = "advance"
        elif billed > paid:
            status = "due"
        else:
            status = "paid"
        v = {
            "name": name,
            "type": old.get("type", ""),
            "billed": billed,
            "paid": paid,
            "status": status,
            "tx": tx_by_vendor.get(name, old.get("tx", [])),
        }
        # Brick Company: show advance payment rows in tx for display (design choice)
        if name == BRICK_VENDOR:
            adv = [[p["date"].strftime("%d-%b"), p["desc"], p["amount"]]
                   for p in pays if p["to"] == BRICK_VENDOR]
            v["tx"] = adv + v["tx"]
        vendors.append(v)
    d["vendors"] = vendors

    # --- categories (with subs from Excel sub-category) ---
    cat_amounts = defaultdict(int)
    sub_amounts = defaultdict(lambda: defaultdict(int))
    for t in txs:
        cat = CATEGORY_MAP.get(t["category"], t["category"])
        cat_amounts[cat] += t["amount"]
        sub_amounts[cat][t["sub"]] += t["amount"]
    old_cats = {c["name"]: c for c in d_old["categories"]}
    cats = []
    ordered_cats = [n for n in old_cats if n in cat_amounts] + \
                   [n for n in cat_amounts if n not in old_cats]
    for name in ordered_cats:
        old = old_cats.get(name, {})
        subs = sorted(sub_amounts[name].items(), key=lambda x: -x[1])
        cats.append({
            "name": name,
            "color": old.get("color", "#8884d8"),
            "amount": cat_amounts[name],
            "subs": [[s, a] for s, a in subs],
        })
    d["categories"] = cats

    # --- monthly ---
    monthly = defaultdict(int)
    for t in txs:
        monthly[t["date"].strftime("%b-%Y")] += t["amount"]
    d["monthly"] = [{"month": m, "billed": a} for m, a in
                    sorted(monthly.items(), key=lambda x: datetime.strptime(x[0], "%b-%Y"))]

    # --- payers ---
    payer_amounts = defaultdict(int)
    for p in pays:
        payer = PAYER_MAP.get(p["by"], p["by"])
        payer_amounts[payer] += p["amount"]
    old_payers = {p["name"]: p for p in d_old["payers"]}
    d["payers"] = [{
        "name": n,
        "sub": old_payers.get(n, {}).get("sub", ""),
        "amount": a,
        "pct": round(a / total_paid * 100, 1),
        "color": old_payers.get(n, {}).get("color", "#8884d8"),
    } for n, a in sorted(payer_amounts.items(), key=lambda x: -x[1])]

    # --- payments list ---
    d["payments"] = [{
        "d": p["date"].strftime("%d-%b"),
        "m": p["date"].strftime("%b-%Y"),
        "by": PAYER_MAP.get(p["by"], p["by"]),
        "to": p["to"],
        "a": p["amount"],
        "ds": p["desc"],
        "md": p["mode"],
    } for p in pays]

    # --- unpaid ---
    last_date = max(p["date"] for p in pays).strftime("%d-%b")
    unpaid = []
    for v in vendors:
        if v["name"] == BRICK_VENDOR:
            continue
        due = v["billed"] - v["paid"]
        if due > 0:
            old_u = next((u for u in d_old.get("unpaid", []) if u["to"] == v["name"]), {})
            unpaid.append({"d": old_u.get("d", last_date), "to": v["name"],
                           "a": due, "ds": old_u.get("ds", "Balance due")})
    d["unpaid"] = unpaid

    # --- summary ---
    balance_due = sum(u["a"] for u in unpaid)
    d["summary"].update({
        "totalBilled": total_billed,
        "totalPaid": total_paid,
        "paymentsCount": len(pays),
        "entriesCount": len(txs),
        "vendorsCount": len(vendors),
        "balanceDue": balance_due,
    })

    # brick advance derived fields
    brick = next((v for v in vendors if v["name"] == BRICK_VENDOR), None)
    if brick:
        d["brickAdvance"]["advancePaid"] = brick["paid"]
        d["brickAdvance"]["deliveredValue"] = brick["billed"]
        d["brickAdvance"]["creditLeft"] = brick["paid"] - brick["billed"]

    # meta
    start = datetime.strptime(d["meta"]["startDate"], "%d-%b-%Y")
    last = max(max(t["date"] for t in txs), max(p["date"] for p in pays))
    d["meta"]["lastUpdated"] = last.strftime("%d-%b-%Y")
    d["meta"]["day"] = (last - start).days + 1
    m = re.match(r"v(\d+)", d_old["meta"].get("version", "v0"))
    d["meta"]["version"] = f"v{int(m.group(1)) + 1}" if m else "v1"
    d["meta"]["validationStatus"] = (
        f"PASSED — vendors, categories & monthly all reconcile to ₹{total_billed:,}".replace(",", ",")
    )
    return d


def reconcile(d):
    errs = []
    s = d["summary"]
    tb, tp, bd = s["totalBilled"], s["totalPaid"], s["balanceDue"]
    checks = [
        ("vendor billed sum", sum(v["billed"] for v in d["vendors"]), tb),
        ("category sum", sum(c["amount"] for c in d["categories"]), tb),
        ("monthly sum", sum(m["billed"] for m in d["monthly"]), tb),
        ("payer sum", sum(p["amount"] for p in d["payers"]), tp),
        ("payments list sum", sum(p["a"] for p in d["payments"]), tp),
        ("unpaid sum", sum(u["a"] for u in d["unpaid"]), bd),
        ("gap", tp - tb, d["brickAdvance"]["creditLeft"] - bd),
        ("vendor paid sum", sum(v["paid"] for v in d["vendors"]), tp),
    ]
    for name, got, want in checks:
        if got != want:
            errs.append(f"{name}: {got:,} != {want:,}")
    for c in d["categories"]:
        ss = sum(a for _, a in c["subs"])
        if ss != c["amount"]:
            errs.append(f"category '{c['name']}' subs {ss:,} != {c['amount']:,}")
    for v in d["vendors"]:
        if v["name"] == BRICK_VENDOR:
            continue
        ts = sum(t[2] for t in v["tx"])
        if ts != v["billed"]:
            errs.append(f"vendor '{v['name']}' tx sum {ts:,} != billed {v['billed']:,}")
    return errs


def sync_html(d):
    html = HTML_PATH.read_text()
    payload = json.dumps(d, ensure_ascii=False, indent=2)
    new_html, n = re.subn(
        r'(<script id="site-data" type="application/json">)(.*?)(</script>)',
        lambda m: m.group(1) + payload + m.group(3), html, flags=re.DOTALL)
    if n != 1:
        raise SystemExit("ERROR: site-data script block not found in index.html")
    HTML_PATH.write_text(new_html)


def main():
    check_only = "--check" in sys.argv
    d_old = json.loads(JSON_PATH.read_text())
    wb = load_workbook(XLSX, data_only=True)
    txs = read_transactions(wb)
    pays = read_payments(wb)
    d = build(d_old, txs, pays)
    errs = reconcile(d)
    s = d["summary"]
    print(f"billed ₹{s['totalBilled']:,} | paid ₹{s['totalPaid']:,} | due ₹{s['balanceDue']:,} "
          f"| {s['entriesCount']} entries | {s['paymentsCount']} payments | {s['vendorsCount']} vendors")
    if errs:
        print("RECONCILIATION FAILED:")
        for e in errs:
            print("  -", e)
        sys.exit(1)
    print("Reconciliation: ALL CHECKS PASSED")
    if check_only:
        print("(--check: no files written)")
        return
    JSON_PATH.write_text(json.dumps(d, ensure_ascii=False, indent=2))
    sync_html(d)
    print(f"Wrote dashboard_data.json ({d['meta']['version']}) and synced index.html")


if __name__ == "__main__":
    main()
